from sec_edgar_api import EdgarClient
import json
from datetime import datetime
from utils.util import filter_frame, get_quarter_values, sorted_by_property, get_sorted_year_report, \
    merge_quarter_result, quarter_type_operatingIncomeLoss, quarter_type_incomeBeforeIncomeTax, \
    quarter_type_incomeBeforeIncomeTaxAndEquityInterest, quarter_type_incomeTaxExpenseBenefit
from openpyxl import load_workbook

import utils.Utils as utils
from openpyxl import Workbook, worksheet


def get_report_summary(wb: Workbook, cik: str, company_name: str):
    edgar = EdgarClient("<Sample Company Name> <Admin Contact>@<Sample Company Domain>")

    # 723125
    # result = edgar.get_company_facts(cik="6951")

    result = edgar.get_company_facts(cik=cik)

    operatingIncomeLossList = result['facts']['us-gaap']['OperatingIncomeLoss']["units"]["USD"]
    incomeTaxExpenseBenefitList = result['facts']['us-gaap']['IncomeTaxExpenseBenefit']["units"]["USD"]

    incomeBeforeIncomeTaxList = None
    if 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest' in result['facts'][
        'us-gaap']:
        incomeBeforeIncomeTaxList = result['facts']['us-gaap'][
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest']["units"][
            "USD"]

    incomeBeforeIncomeTaxAndEquityInterestList = None
    if 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments' in \
            result['facts']['us-gaap']:
        incomeBeforeIncomeTaxAndEquityInterestList = result['facts']['us-gaap'][
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'][
            "units"]["USD"]

    # sort the list
    operatingIncomeLossList = sorted_by_property(operatingIncomeLossList, "end")
    incomeTaxExpenseBenefitList = sorted_by_property(incomeTaxExpenseBenefitList, "end")
    incomeBeforeIncomeTaxList = sorted_by_property(incomeBeforeIncomeTaxList, "end")
    incomeBeforeIncomeTaxAndEquityInterestList = sorted_by_property(incomeBeforeIncomeTaxAndEquityInterestList, "end")

    # filter data with "frame" property
    operatingIncomeLossList = filter_frame(operatingIncomeLossList)
    incomeTaxExpenseBenefitList = filter_frame(incomeTaxExpenseBenefitList)
    incomeBeforeIncomeTaxList = filter_frame(incomeBeforeIncomeTaxList)
    incomeBeforeIncomeTaxAndEquityInterestList = filter_frame(incomeBeforeIncomeTaxAndEquityInterestList)

    sheet = wb.create_sheet(title=company_name)
    keys = operatingIncomeLossList[0].keys()

    for index, key in enumerate(keys):
        cell = sheet.cell(row=1, column=index + 1)
        cell.value = key

    for index, data in enumerate(operatingIncomeLossList):
        # print(data)
        op_income_loss = operatingIncomeLossList[index]
        income_tax_expense = None
        income_before_income_tax = None
        income_before_income_tax_equity = None

        if index < len(incomeTaxExpenseBenefitList):
            income_tax_expense = incomeTaxExpenseBenefitList[index]

        if index < len(incomeBeforeIncomeTaxList):
            income_before_income_tax = incomeBeforeIncomeTaxList[index]

        if index < len(incomeBeforeIncomeTaxAndEquityInterestList):
            income_before_income_tax_equity = incomeBeforeIncomeTaxAndEquityInterestList[index]

        for col_index, key in enumerate(op_income_loss.keys()):
            cell = sheet.cell(row=1+index, column=1+col_index)
            cell.value = ""
        cell = sheet.cell(row=1 + index, column=0)
        # start_date = data['start']
        # end_date = data['end']
        # value = data['val']
        # fp_year = data['fp']


output_folder = "output/"
result_file = output_folder + "result.xlsx"
utils.delete_directory(output_folder)
utils.create_directory(output_folder)

company_dict = {"MU": "723125", "AVGO": "1730168", "ADI": "6281", "TXN": "97476", "MRVL": "1835632", "SNPS": "883241",
                "CDNS": "813672", "DELL": "1571996", "AAPL": "320193", "HP": "46765"}

company_dict = {"MU": "723125"}

wb = Workbook()

for name in company_dict:
    cik = company_dict[name]

    get_report_summary(wb=wb, cik=cik, company_name=name)

wb.save(filename=result_file)
