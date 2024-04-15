from sec_edgar_api import EdgarClient
import json
from datetime import datetime
from utils.util import filter_frame, sorted_by_property, merge_property_value, get_simple_quarter, toMillion, \
    QuarterSimple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

import utils.Utils as utils
from openpyxl import Workbook, worksheet
from openpyxl.styles import PatternFill, Alignment, Font, Color


def get_report_summary(wb: Workbook, cik: str, company_name: str):
    edgar = EdgarClient("<Sample Company Name> <Admin Contact>@<Sample Company Domain>")

    # 723125
    # result = edgar.get_company_facts(cik="6951")

    result = edgar.get_company_facts(cik=cik)

    operatingIncomeLossList = result['facts']['us-gaap']['OperatingIncomeLoss']["units"]["USD"]
    incomeTaxExpenseBenefitList = result['facts']['us-gaap']['IncomeTaxExpenseBenefit']["units"]["USD"]

    incomeBeforeIncomeTaxList = None
    if 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest' in result['facts'][
        'us-gaap']:
        incomeBeforeIncomeTaxList = result['facts']['us-gaap'][
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest']["units"][
            "USD"]

    incomeBeforeIncomeTaxAndEquityInterestList = None
    if 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments' in \
            result['facts']['us-gaap']:
        incomeBeforeIncomeTaxAndEquityInterestList = result['facts']['us-gaap'][
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'][
            "units"]["USD"]

    # sort the list
    operatingIncomeLossList = sorted_by_property(operatingIncomeLossList, "end")

    # filter data with "frame" property
    operatingIncomeLossList = filter_frame(operatingIncomeLossList)

    row_names = [
        "Fiscal Quarter",
        "Calendar Quarter",
        "start",
        "end",
        "Operating Income Loss",
        "Income Tax Expense Benefit",
        "Income Before Income Tax",
        "Income Before Income Tax And EquityInterest"]

    # consolidate property
    merge_property_value(toList=operatingIncomeLossList, fromList=operatingIncomeLossList,
                         name=row_names[4])
    merge_property_value(toList=operatingIncomeLossList, fromList=incomeTaxExpenseBenefitList,
                         name=row_names[5])
    merge_property_value(toList=operatingIncomeLossList, fromList=incomeBeforeIncomeTaxList,
                         name=row_names[6])
    merge_property_value(toList=operatingIncomeLossList, fromList=incomeBeforeIncomeTaxAndEquityInterestList,
                         name=row_names[7])

    # operatingIncomeLossList.reverse()
    # res_list = []
    #
    # for qDict in operatingIncomeLossList:
    #     print(qDict)
    #     q = get_simple_quarter(qDict)
    #
    #     if q.form == "10-Q":
    #         res_list.insert(0, qDict)
    #     elif q.form == "10-K":
    #         last_q = {}


    sheet = wb.create_sheet(title=company_name)

    # write the row
    for row_index, row_name in enumerate(row_names):
        cell = sheet.cell(row=row_index + 1, column=1)
        cell.value = row_name
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = Font(size="12", bold=True)
        # cell.fill = PatternFill(start_color='FFCCE5FF',
        #                         end_color='FFCCE5FF',
        #                         fill_type='solid')
        sheet.row_dimensions[row_index + 1].height = 40
    sheet.column_dimensions[get_column_letter(1)].width = 40

    for column_index, qDict in enumerate(operatingIncomeLossList):
        sheet.column_dimensions[get_column_letter(column_index + 2)].width = 15
        q = get_simple_quarter(qDict=qDict)
        values = [q.fp, q.frame, q.start, q.end, q.operatingIncomeLoss, q.incomeTaxExpenseBenefit,
                  q.incomeBeforeIncomeTax, q.incomeBeforeIncomeTaxAndEquityInterest]

        for row_index, row_name in enumerate(row_names):
            cell = sheet.cell(row=row_index + 1, column=column_index + 2)
            cell.value = values[row_index]
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size="12", color=Color())
            # cell.fill = PatternFill(start_color='FF99CCFF',
            #                         end_color='FF99CCFF',
            #                         fill_type='solid')


output_folder = "output/"
result_file = output_folder + "result.xlsx"
utils.delete_directory(output_folder)
utils.create_directory(output_folder)

company_dict = {"MU": "723125", "AVGO": "1730168", "ADI": "6281", "TXN": "97476", "MRVL": "1835632", "SNPS": "883241",
                "CDNS": "813672", "DELL": "1571996", "AAPL": "320193", "HP": "46765"}

# company_dict = {"MU": "723125"}

wb = Workbook()

for name in company_dict:
    cik = company_dict[name]

    get_report_summary(wb=wb, cik=cik, company_name=name)

del wb["Sheet"]
wb.save(filename=result_file)
