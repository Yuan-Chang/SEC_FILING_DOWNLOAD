import os
from openpyxl import load_workbook, Workbook
from pathlib import Path
import shutil
from typing import List
from copy import copy


def create_values_only_excel_file(input_file, output_file):
    wb = load_workbook(input_file, data_only=True)
    wb.save(output_file)


def is_file_exist(file_path):
    file = Path(file_path)
    if file.is_file():
        return True
    else:
        return False


def delete_file(file_path):
    file = Path(file_path)

    if file.is_file():
        file.unlink()


def create_file_copy(input_file, output_file):
    shutil.copyfile(input_file, output_file)


def create_directory(directory_path):
    os.mkdir(directory_path)


def delete_directory(directory_path):
    file = Path(directory_path)

    if file.is_dir():
        shutil.rmtree(directory_path)


def get_file_name_from_file_path(file_path):
    return Path(file_path).name


def get_file_name_without_extension_from_file_path(file_path):
    return Path(file_path).stem


def quit_excel():
    os.system('taskkill /T /IM EXCEL.exe')


def create_sheet(excel_file_path, sheet_name):
    wb = load_workbook(excel_file_path)
    wb.create_sheet(sheet_name)
    wb.save(excel_file_path)


def copy_worksheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws[cell.coordinate].value = cell.value
            target_ws[cell.coordinate].font = copy(cell.font)
            target_ws[cell.coordinate].border = copy(cell.border)
            target_ws[cell.coordinate].fill = copy(cell.fill)
            target_ws[cell.coordinate].number_format = cell.number_format
            target_ws[cell.coordinate].protection = copy(cell.protection)
            target_ws[cell.coordinate].alignment = copy(cell.alignment)
            target_ws[cell.coordinate].comment = cell.comment

    # Copy cell width and height
    for idx, rd in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx] = copy(rd)

    for idx, rd in source_ws.column_dimensions.items():
        target_ws.column_dimensions[idx] = copy(rd)

    for merged_cell in source_ws.merged_cells:
        target_ws.merge_cells(f"{merged_cell}")


def merge_multiple_excels_to_one_excel(input_files: List, output_file):
    output_wb = Workbook()

    for file_name in input_files:
        source_wb = load_workbook(file_name)
        number_of_sheets = len(source_wb.sheetnames)

        for sheet in source_wb:
            file_name = get_file_name_without_extension_from_file_path(file_name)
            sheet_name = f"{file_name}_{sheet.title}"

            # if only 1 sheet, use file name instead
            if number_of_sheets == 1:
                sheet_name = file_name

            output_ws = output_wb.create_sheet(sheet_name)
            copy_worksheet(sheet, output_ws)

    # Delete the default sheet
    del output_wb["Sheet"]
    output_wb.save(output_file)


def replace_extension(file_name, extension):
    return file_name.rsplit('.', 1)[0] + "." + extension
